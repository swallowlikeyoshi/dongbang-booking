"""명부 xlsx + 스터디 시트 xlsx → data/roster.csv (student_no,name,sub_team).

실행:
  uvx --from openpyxl python scripts/roster-to-csv.py <명부.xlsx> <스터디시트.xlsx>
"""
import sys, csv, pathlib
import openpyxl

roster_path, study_path = sys.argv[1], sys.argv[2]

wb = openpyxl.load_workbook(roster_path, data_only=True)
rows = [r for r in wb["시트1"].iter_rows(min_row=4, values_only=True)
        if r[0] and str(r[0]).strip() != "이름"]
# 학번 기준 dedupe. 임원이 직책 행과 정회원 행으로 2회 등재되어 있다.
elec = {}
for r in rows:
    if str(r[4]).strip() == "전기":
        elec[str(r[0]).strip()] = str(r[3]).replace(".0", "")

wb2 = openpyxl.load_workbook(study_path, data_only=True)
study = [(str(r[0]).strip(), r[1]) for r in wb2["총계"].iter_rows(min_row=3, values_only=True) if r[0]]

out = []
for name, team in study:
    if not team:
        continue          # 세부팀 미배정 제외
    if name not in elec:
        continue          # 명부에 없으면 제외
    out.append((elec[name], name, str(team).strip()))

pathlib.Path("data").mkdir(exist_ok=True)
with open("data/roster.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["student_no", "name", "sub_team"])
    w.writerows(out)
print(f"wrote data/roster.csv: {len(out)} members")
